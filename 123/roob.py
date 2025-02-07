import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
import os


def save_to_excel():
    # Проверка обязательных полей
    if not (entry_company_name.get() and entry_company_address.get() and
            entry_commission_chair.get() and entry_order_number.get() and
            entry_opr_date.get() and
            entry_member1.get() and entry_member2.get() and entry_member3.get()):
        messagebox.showerror("Ошибка", "Пожалуйста, заполните все обязательные поля.")
        return

    # Сбор данных
    data = {
        "Наименование компании": entry_company_name.get(),
        "Адрес компании": entry_company_address.get(),
        "Председатель комиссии": entry_commission_chair.get(),
        "Номер и дата приказа": entry_order_number.get(),
        "Дата проведения ОПР": entry_opr_date.get(),
        "Член комиссии 1": entry_member1.get(),
        "Член комиссии 2": entry_member2.get(),
        "Член комиссии 3": entry_member3.get(),
        "Член комиссии 4": entry_member4.get(),
        "Член комиссии 5": entry_member5.get()
    }

    file_name = "user_data.xlsx"

    # Проверка: существует ли файл Excel
    if not os.path.exists(file_name):
        # Создаем новый Excel-файл
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Данные"

        # Добавляем заголовки
        headers = list(data.keys())
        sheet.append(headers)  # Записываем заголовки в первую строку

        # Сохраняем файл
        workbook.save(file_name)

    # Открываем существующий Excel-файл и добавляем новую строку данных
    workbook = load_workbook(file_name)
    sheet = workbook.active

    # Добавляем новую строку с данными
    new_row = [data[key] for key in data.keys()]
    sheet.append(new_row)

    try:
        workbook.save(file_name)
        messagebox.showinfo("Успех", "Данные успешно сохранены в Excel!")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить данные: {e}")


# Создание главного окна
root = tk.Tk()
root.title("Форма для ввода данных")

# Поля формы
tk.Label(root, text="Наименование компании:").grid(row=0, column=0, padx=10, pady=5)
entry_company_name = tk.Entry(root, width=50)
entry_company_name.grid(row=0, column=1, padx=10, pady=5)

tk.Label(root, text="Адрес компании:").grid(row=1, column=0, padx=10, pady=5)
entry_company_address = tk.Entry(root, width=50)
entry_company_address.grid(row=1, column=1, padx=10, pady=5)

tk.Label(root, text="Председатель комиссии:").grid(row=2, column=0, padx=10, pady=5)
entry_commission_chair = tk.Entry(root, width=50)
entry_commission_chair.grid(row=2, column=1, padx=10, pady=5)

tk.Label(root, text="Номер и дата приказа:").grid(row=3, column=0, padx=10, pady=5)
entry_order_number = tk.Entry(root, width=50)
entry_order_number.grid(row=3, column=1, padx=10, pady=5)


tk.Label(root, text="Дата проведения ОПР:").grid(row=5, column=0, padx=10, pady=5)
entry_opr_date = tk.Entry(root, width=50)
entry_opr_date.grid(row=5, column=1, padx=10, pady=5)

# Поля для членов комиссии
tk.Label(root, text="Член комиссии 1 (обязательно):").grid(row=6, column=0, padx=10, pady=5)
entry_member1 = tk.Entry(root, width=50)
entry_member1.grid(row=6, column=1, padx=10, pady=5)

tk.Label(root, text="Член комиссии 2 (обязательно):").grid(row=7, column=0, padx=10, pady=5)
entry_member2 = tk.Entry(root, width=50)
entry_member2.grid(row=7, column=1, padx=10, pady=5)

tk.Label(root, text="Член комиссии 3 (обязательно):").grid(row=8, column=0, padx=10, pady=5)
entry_member3 = tk.Entry(root, width=50)
entry_member3.grid(row=8, column=1, padx=10, pady=5)

tk.Label(root, text="Член комиссии 4 (необязательно):").grid(row=9, column=0, padx=10, pady=5)
entry_member4 = tk.Entry(root, width=50)
entry_member4.grid(row=9, column=1, padx=10, pady=5)

tk.Label(root, text="Член комиссии 5 (необязательно):").grid(row=10, column=0, padx=10, pady=5)
entry_member5 = tk.Entry(root, width=50)
entry_member5.grid(row=10, column=1, padx=10, pady=5)

# Кнопка для сохранения данных
save_button = tk.Button(root, text="Сохранить данные", command=save_to_excel)
save_button.grid(row=11, columnspan=2, pady=10)

# Запуск приложения
root.mainloop()